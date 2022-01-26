import os
from datetime import datetime
from openpyxl import Workbook
from packages import SubmissionInformationPackage

"""
This script looks for ETDs that have been cataloged.
It then adds the Bib record ID to the SIP (and possibly the ProQuest ID?).
Finally it creates a derivative package for ingest into the IR.
"""

def initializeWorkbook():
    wb = Workbook()
    ws = wb.active
    headings = ["title", "fulltext_url", "keywords", "abstract", "author1_fname", "author1_mname", "author1_lname",\
     "author1_suffix", "author1_email", "author1_institution", "advisor1", "advisor2", "advisor3", "disciplines",\
     "comments", "degree_name", "department", "document_type", "embargo_date", "publication_date", "season"]
    ws.append(headings)
    
    return wb;

#version of readCatalogedSIPs.py
version = "0.1"
if os.name == "nt":
    catalogingPath = "\\\\Lincoln\\Library\\ETDs\\CatalogingPackage"
    irPath = "\\\\Lincoln\\Library\\ETDs\\IRPackages\\incoming"
else:
    catalogingPath = "/media/Library/ETDs/CatalogingPackage"
    irPath = "/media/Library/ETDs/IRPackages/incoming"
catalogingIncomingPath = os.path.join(catalogingPath, "outgoing")

#ingestSpreadsheet = os.path.join(irPath, "ingest_" + datetime.now().strftime("%d-%m-%Y_%H-%M-%S") + ".xlsx")

#initialize hashmap, workbook per department
hashmap_wb = dict()

#wb = Workbook()
#ws = wb.active
#headings = ["title", "fulltext_url", "keywords", "abstract", "author1_fname", "author1_mname", "author1_lname",\
# "author1_suffix", "author1_email", "author1_institution", "advisor1", "advisor2", "advisor3", "disciplines",\
# "comments", "degree_name", "department", "document_type", "embargo_date", "publication_date", "season"]
#ws.append(headings)

for package in os.listdir(catalogingIncomingPath):
    print ("Reading " + package + "...")

    SIP = SubmissionInformationPackage()
    pathSIP = SIP.lookup(os.path.join(catalogingIncomingPath, package))
    SIP.load(pathSIP)
    #print (SIP.bag.is_valid())

    print ("\t" + SIP.identifier)
    print ("\t" + SIP.bag.info['Author-1_lname'])
    SIP.bag.info['Bib-Record'] = package
    #SIP.bag.save()
    #print (SIP.bag.is_valid())
    
    # build the spreadsheet row
    row = []
    row = SIP.writeField(row, "Title")
    row.append("/path/to/webserver")
    row = SIP.writeField(row, "Keywords")
    row = SIP.writeField(row, "Abstract")
    row = SIP.writeField(row, "Author-1_fname")
    row = SIP.writeField(row, "Author-1_mname")
    row = SIP.writeField(row, "Author-1_lname")
    row = SIP.writeField(row, "Author-1_suffix")
    row = SIP.writeField(row, "Author-1_email_future", "Author-1_email_current")
    row = SIP.writeField(row, "Institution")
    row = SIP.writeField(row, "advisor1")
    row = SIP.writeField(row, "advisor2")
    row = SIP.writeField(row, "advisor3")
    row = SIP.writeField(row, "disciplines")
    #for comments cell
    row.append("")
    row = SIP.writeField(row, "Degree-Name")
    row = SIP.writeField(row, "Department")
    #for document_type, embargo_date, publication_date, and season
    row.append("")
    row.append("")
    row.append("")
    row.append("")
    
    # identify department, call up associated Workbook and worksheet
    dept = SIP.bag.info['Department']
    
    if not dept in hashmap_wb:
        hashmap_wb[dept] = initializeWorkbook()
    
    wb = hashmap_wb[dept]
    ws = wb.active
    ws.append(row)
    
    # remove cataloging package
    print ("Removing cataloging package " + package)
    #os.remove(os.path.join(incomingPath, package))

    # Build IR package
    #SIP.makeIRPackage()

##write IR ingest spreadsheet
#wb.save(ingestSpreadsheet)

#write IR injest spreadsheets per department
depts = hashmap_wb.keys()
for dept in depts:
    wb = hashmap_wb[dept]
    wb.save(os.path.join(irPath, "ingest_" + dept + "_" + datetime.now().strftime("%d-%m-%Y_%H-%M-%S") + ".xlsx"))
    